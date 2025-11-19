#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Teste de Exportação Excel - Demonstra os tipos de exportação disponíveis
"""

from models.database import Database
from utils.excel_export import gerar_excel_relatorio_completo, exportar_apenas_emprestimos
from pathlib import Path
import openpyxl

def testar_exportacoes():
    print("=" * 70)
    print("TESTE: Exportação para Excel")
    print("=" * 70)
    
    db = Database()
    
    # Teste 1: Relatório Completo
    print("\n1. Gerando Relatório Completo (4 abas)...")
    try:
        caminho1 = gerar_excel_relatorio_completo(db, "teste_relatorio_completo.xlsx")
        print(f"   ✓ Criado: {caminho1}")
        
        # Verificar abas
        wb = openpyxl.load_workbook(caminho1)
        abas = wb.sheetnames
        print(f"   ✓ Abas criadas: {', '.join(abas)}")
        
        # Contar linhas
        for aba in abas:
            ws = wb[aba]
            linhas = ws.max_row
            print(f"      - {aba}: {linhas - 1} registros")
        
        wb.close()
        
    except Exception as e:
        print(f"   ✗ Erro: {e}")
    
    # Teste 2: Apenas Empréstimos
    print("\n2. Gerando Relatório de Empréstimos...")
    try:
        caminho2 = exportar_apenas_emprestimos(db, "teste_emprestimos.xlsx")
        print(f"   ✓ Criado: {caminho2}")
        
        # Verificar conteúdo
        wb = openpyxl.load_workbook(caminho2)
        ws = wb.active
        
        linhas_dados = ws.max_row - 1
        colunas = ws.max_column
        
        print(f"   ✓ Empréstimos: {linhas_dados}")
        print(f"   ✓ Colunas: {colunas}")
        
        wb.close()
        
    except Exception as e:
        print(f"   ✗ Erro: {e}")
    
    # Teste 3: Verificar tamanhos
    print("\n3. Tamanhos dos arquivos:")
    for arquivo in ["teste_relatorio_completo.xlsx", "teste_emprestimos.xlsx"]:
        try:
            tamanho_kb = Path(arquivo).stat().st_size / 1024
            print(f"   📁 {arquivo}: {tamanho_kb:.1f} KB")
        except:
            pass
    
    print("\n" + "=" * 70)
    print("✓ Testes de Exportação Concluídos!")
    print("=" * 70)

if __name__ == "__main__":
    testar_exportacoes()
