#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Teste de Exportação Excel com Dados Reais
"""

from models.database import Database
from models.cliente import Cliente
from models.emprestimo import Emprestimo
from utils.excel_export import gerar_excel_relatorio_completo, exportar_apenas_emprestimos
from pathlib import Path
import openpyxl

def testar_exportacao_com_dados():
    print("=" * 70)
    print("TESTE: Exportação Excel com Dados Reais")
    print("=" * 70)
    
    # Criar database temporária com dados
    db = Database()
    
    # Criar clientes de teste
    print("\n1. Adicionando clientes de teste...")
    clientes = [
        Cliente("João Silva", "123.456.789-00", "11999999999", "joao@email.com", "Rua A, 123"),
        Cliente("Maria Santos", "987.654.321-00", "11888888888", "maria@email.com", "Rua B, 456"),
        Cliente("Pedro Costa", "456.123.789-00", "11777777777", "pedro@email.com", "Rua C, 789"),
    ]
    
    for cliente in clientes:
        db.clientes.append(cliente)
    
    print(f"   ✓ {len(clientes)} clientes adicionados")
    
    # Criar empréstimos de teste
    print("\n2. Adicionando empréstimos de teste...")
    
    # Ativo com pagamento
    emp1 = Emprestimo(
        cliente_id=clientes[0].id,
        valor_emprestado=5000.0,
        taxa_juros=2.5,
        data_inicio="2024-09-01",
        prazo_meses=12
    )
    emp1.registrar_pagamento(500.0, data="2024-10-01")
    emp1.registrar_pagamento(500.0, data="2024-11-01")
    db.emprestimos.append(emp1)
    
    # Ativo sem pagamento
    emp2 = Emprestimo(
        cliente_id=clientes[1].id,
        valor_emprestado=10000.0,
        taxa_juros=3.0,
        data_inicio="2024-10-15",
        prazo_meses=24
    )
    db.emprestimos.append(emp2)
    
    # Quitado
    emp3 = Emprestimo(
        cliente_id=clientes[2].id,
        valor_emprestado=3000.0,
        taxa_juros=2.0,
        data_inicio="2024-01-01",
        prazo_meses=6
    )
    # Simular quitação
    total = emp3.valor_total
    emp3.registrar_pagamento(total)
    db.emprestimos.append(emp3)
    
    print(f"   ✓ {len(db.emprestimos)} empréstimos adicionados")
    print(f"      - 1 Ativo com 2 pagamentos")
    print(f"      - 1 Ativo sem pagamentos")
    print(f"      - 1 Quitado")
    
    # Teste 1: Relatório Completo
    print("\n3. Gerando Relatório Completo com dados...")
    try:
        caminho1 = gerar_excel_relatorio_completo(db, "teste_completo_dados.xlsx")
        print(f"   ✓ Criado: {caminho1}")
        
        # Verificar conteúdo
        wb = openpyxl.load_workbook(caminho1)
        
        # Verificar Resumo
        ws_resumo = wb["Resumo"]
        print(f"\n   RESUMO (primeiras 15 linhas):")
        for row in range(1, min(16, ws_resumo.max_row + 1)):
            cell_a = ws_resumo[f'A{row}'].value
            cell_b = ws_resumo[f'B{row}'].value
            if cell_a:
                print(f"      {cell_a}: {cell_b}")
        
        # Verificar Clientes
        ws_clientes = wb["Clientes"]
        print(f"\n   CLIENTES: {ws_clientes.max_row - 1} registros")
        for row in range(2, ws_clientes.max_row + 1):
            nome = ws_clientes[f'B{row}'].value
            if nome:
                print(f"      - {nome}")
        
        # Verificar Empréstimos
        ws_emps = wb["Empréstimos"]
        print(f"\n   EMPRÉSTIMOS: {ws_emps.max_row - 1} registros")
        for row in range(2, ws_emps.max_row + 1):
            id_emp = ws_emps[f'A{row}'].value
            cliente = ws_emps[f'B{row}'].value
            status = ws_emps[f'K{row}'].value
            saldo = ws_emps[f'G{row}'].value
            if id_emp:
                print(f"      - {cliente} ({status}): Saldo R$ {saldo:.2f}")
        
        # Verificar Pagamentos
        ws_pgs = wb["Pagamentos"]
        print(f"\n   PAGAMENTOS: {ws_pgs.max_row - 1} registros")
        for row in range(2, ws_pgs.max_row + 1):
            id_pgt = ws_pgs[f'A{row}'].value
            cliente = ws_pgs[f'C{row}'].value
            valor = ws_pgs[f'D{row}'].value
            if id_pgt:
                print(f"      - {cliente}: R$ {valor:.2f}")
        
        wb.close()
        
    except Exception as e:
        print(f"   ✗ Erro: {e}")
        import traceback
        traceback.print_exc()
    
    # Teste 2: Apenas Empréstimos
    print("\n4. Gerando Relatório de Empréstimos...")
    try:
        caminho2 = exportar_apenas_emprestimos(db, "teste_emps_dados.xlsx")
        print(f"   ✓ Criado: {caminho2}")
        
        tamanho_kb = Path(caminho2).stat().st_size / 1024
        print(f"   📁 Tamanho: {tamanho_kb:.1f} KB")
        
    except Exception as e:
        print(f"   ✗ Erro: {e}")
    
    print("\n" + "=" * 70)
    print("✓ Testes de Exportação Com Dados Concluídos!")
    print("=" * 70)

if __name__ == "__main__":
    testar_exportacao_com_dados()
