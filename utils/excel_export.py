# -*- coding: utf-8 -*-
"""
Exportação para Excel - Gera planilhas bem formatadas com dados de clientes e empréstimos
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path


def formatar_moeda(valor):
    """Formata valor como moeda brasileira"""
    return f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')


def formatar_data_br(data_iso):
    """Converte data ISO para formato BR DD/MM/YYYY"""
    if not data_iso:
        return ""
    try:
        data_obj = datetime.fromisoformat(data_iso)
        return data_obj.strftime("%d/%m/%Y")
    except:
        return data_iso


class EstiloExcel:
    """Estilos para as planilhas"""
    
    # Cores
    COR_HEADER = "1F4E78"  # Azul escuro
    COR_TITULO = "4472C4"  # Azul
    COR_TOTAL = "D9E1F2"   # Azul claro
    COR_QUITADO = "E2EFDA" # Verde claro
    COR_ATIVO = "FFF2CC"   # Amarelo claro
    COR_ATRASADO = "F4CCCC" # Vermelho claro
    
    # Fontes
    FONTE_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    FONTE_TITULO = Font(name="Calibri", size=14, bold=True, color=COR_HEADER)
    FONTE_BOLD = Font(name="Calibri", size=11, bold=True)
    FONTE_NORMAL = Font(name="Calibri", size=11)
    
    # Preenchimentos
    FILL_HEADER = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    FILL_TOTAL = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    FILL_QUITADO = PatternFill(start_color=COR_QUITADO, end_color=COR_QUITADO, fill_type="solid")
    FILL_ATIVO = PatternFill(start_color=COR_ATIVO, end_color=COR_ATIVO, fill_type="solid")
    FILL_ATRASADO = PatternFill(start_color=COR_ATRASADO, end_color=COR_ATRASADO, fill_type="solid")
    
    # Bordas
    BORDA = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @staticmethod
    def aplicar_header(cell, merge=False):
        """Aplica estilo de header a uma célula"""
        cell.font = EstiloExcel.FONTE_HEADER
        cell.fill = EstiloExcel.FILL_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = EstiloExcel.BORDA
    
    @staticmethod
    def aplicar_normal(cell):
        """Aplica estilo normal a uma célula"""
        cell.font = EstiloExcel.FONTE_NORMAL
        cell.border = EstiloExcel.BORDA
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    @staticmethod
    def aplicar_numerico(cell, valor=None):
        """Aplica estilo numérico"""
        cell.font = EstiloExcel.FONTE_NORMAL
        cell.border = EstiloExcel.BORDA
        cell.alignment = Alignment(horizontal='right', vertical='center')
        if valor is not None and isinstance(valor, (int, float)):
            cell.number_format = '#,##0.00'
    
    @staticmethod
    def aplicar_data(cell):
        """Aplica estilo de data"""
        cell.font = EstiloExcel.FONTE_NORMAL
        cell.border = EstiloExcel.BORDA
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    @staticmethod
    def aplicar_status(cell, status):
        """Aplica estilo baseado em status"""
        if status == "Quitado":
            cell.fill = EstiloExcel.FILL_QUITADO
        elif status == "Ativo":
            cell.fill = EstiloExcel.FILL_ATIVO
        elif status == "Atrasado":
            cell.fill = EstiloExcel.FILL_ATRASADO
        
        EstiloExcel.aplicar_normal(cell)


def gerar_excel_relatorio_completo(database, caminho_saida=None):
    """
    Gera Excel com múltiplas abas: Resumo, Clientes, Empréstimos, Pagamentos
    """
    if not caminho_saida:
        caminho_saida = f"relatorio_financeiro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove a aba padrão
    
    # Aba 1: Resumo Executivo
    _criar_aba_resumo(wb, database)
    
    # Aba 2: Clientes
    _criar_aba_clientes(wb, database)
    
    # Aba 3: Empréstimos
    _criar_aba_emprestimos(wb, database)
    
    # Aba 4: Pagamentos
    _criar_aba_pagamentos(wb, database)
    
    # Salvar arquivo
    wb.save(caminho_saida)
    return caminho_saida


def _criar_aba_resumo(wb, database):
    """Cria aba de resumo executivo"""
    ws = wb.create_sheet("Resumo", 0)
    
    # Título
    ws.merge_cells('A1:D1')
    cell_titulo = ws['A1']
    cell_titulo.value = "RESUMO EXECUTIVO"
    cell_titulo.font = EstiloExcel.FONTE_TITULO
    cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Data de geração
    ws['A2'] = f"Gerado em: {formatar_data_br(datetime.now().isoformat())}"
    
    # Espaço
    row = 4
    
    # Estatísticas Gerais
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "ESTATÍSTICAS GERAIS"
    cell.font = EstiloExcel.FONTE_BOLD
    cell.fill = EstiloExcel.FILL_TOTAL
    EstiloExcel.aplicar_header(cell)
    row += 1
    
    # Calcular métricas
    total_clientes = len(database.clientes)
    total_emprestimos = len(database.emprestimos)
    emprestimos_ativos = len([e for e in database.emprestimos if e.ativo])
    emprestimos_quitados = len([e for e in database.emprestimos if not e.ativo])
    
    valor_total_emprestado = sum(e.valor_emprestado for e in database.emprestimos)
    valor_total_receber = sum(e.saldo_devedor for e in database.emprestimos)
    valor_total_recebido = valor_total_emprestado - valor_total_receber
    total_juros = sum(e.total_juros for e in database.emprestimos)
    
    # Dados
    dados_resumo = [
        ("Total de Clientes", total_clientes),
        ("Total de Empréstimos", total_emprestimos),
        ("  ├─ Ativos", emprestimos_ativos),
        ("  └─ Quitados", emprestimos_quitados),
        ("", ""),
        ("Valor Total Emprestado", valor_total_emprestado),
        ("Valor Total Já Recebido", valor_total_recebido),
        ("Valor Total a Receber", valor_total_receber),
        ("Total de Juros", total_juros),
    ]
    
    for label, valor in dados_resumo:
        ws[f'A{row}'] = label
        if label == "":
            row += 1
            continue
        
        if isinstance(valor, (int, float)):
            ws[f'B{row}'] = valor
            if valor > 0 or label.startswith("  "):
                EstiloExcel.aplicar_numerico(ws[f'B{row}'], valor)
                if valor > 0:
                    ws[f'B{row}'].number_format = '_("R$"* #,##0.00_);_("R$"* (#,##0.00);_("R$"* "-"??_);_(@_)'
            else:
                EstiloExcel.aplicar_normal(ws[f'B{row}'])
        
        EstiloExcel.aplicar_normal(ws[f'A{row}'])
        row += 1
    
    # Ajustar colunas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def _criar_aba_clientes(wb, database):
    """Cria aba de clientes"""
    ws = wb.create_sheet("Clientes")
    
    # Headers
    headers = ["ID", "Nome", "CPF/CNPJ", "Email", "Telefone", "Endereço", "Chave PIX", "Data Cadastro"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        EstiloExcel.aplicar_header(cell)
    
    ws.row_dimensions[1].height = 20
    
    # Dados
    for row, cliente in enumerate(database.clientes, 2):
        ws.cell(row, 1).value = cliente.id
        ws.cell(row, 2).value = cliente.nome
        ws.cell(row, 3).value = cliente.cpf_cnpj
        ws.cell(row, 4).value = cliente.email
        ws.cell(row, 5).value = cliente.telefone
        ws.cell(row, 6).value = cliente.endereco
        ws.cell(row, 7).value = cliente.chave_pix
        ws.cell(row, 8).value = formatar_data_br(cliente.data_cadastro)
        
        for col in range(1, len(headers) + 1):
            EstiloExcel.aplicar_normal(ws.cell(row, col))
    
    # Ajustar colunas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15


def _criar_aba_emprestimos(wb, database):
    """Cria aba de empréstimos"""
    ws = wb.create_sheet("Empréstimos")
    
    # Headers
    headers = [
        "ID", "Cliente", "Valor Emprestado", "Taxa Juros", "Prazo (meses)",
        "Valor Total", "Saldo Devedor", "Total Pago", "% Pago",
        "Data Início", "Status", "Ativo"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        EstiloExcel.aplicar_header(cell)
    
    ws.row_dimensions[1].height = 20
    
    # Dados
    for row, emprestimo in enumerate(database.emprestimos, 2):
        cliente = database.get_cliente_por_id(emprestimo.cliente_id)
        cliente_nome = cliente.nome if cliente else "N/A"
        
        total_pago = emprestimo.valor_total - emprestimo.saldo_devedor
        percentual_pago = (total_pago / emprestimo.valor_total * 100) if emprestimo.valor_total > 0 else 0
        
        status = "Quitado" if emprestimo.saldo_devedor <= 0 else "Ativo"
        
        ws.cell(row, 1).value = emprestimo.id
        ws.cell(row, 2).value = cliente_nome
        ws.cell(row, 3).value = emprestimo.valor_emprestado
        ws.cell(row, 4).value = emprestimo.taxa_juros * 100
        ws.cell(row, 5).value = emprestimo.prazo_meses
        ws.cell(row, 6).value = emprestimo.valor_total
        ws.cell(row, 7).value = emprestimo.saldo_devedor
        ws.cell(row, 8).value = total_pago
        ws.cell(row, 9).value = percentual_pago
        ws.cell(row, 10).value = formatar_data_br(emprestimo.data_emprestimo)
        ws.cell(row, 11).value = status
        ws.cell(row, 12).value = "Sim" if emprestimo.ativo else "Não"
        
        # Aplicar estilos
        EstiloExcel.aplicar_normal(ws.cell(row, 1))
        EstiloExcel.aplicar_normal(ws.cell(row, 2))
        
        for col in [3, 4, 6, 7, 8]:
            EstiloExcel.aplicar_numerico(ws.cell(row, col), ws.cell(row, col).value)
        
        EstiloExcel.aplicar_numerico(ws.cell(row, 9), percentual_pago)
        ws.cell(row, 9).number_format = '0.00"%"'
        
        EstiloExcel.aplicar_data(ws.cell(row, 10))
        EstiloExcel.aplicar_status(ws.cell(row, 11), status)
        EstiloExcel.aplicar_normal(ws.cell(row, 12))
    
    # Ajustar colunas
    larguras = [15, 20, 16, 12, 14, 14, 15, 14, 10, 14, 12, 10]
    for col, largura in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = largura


def _criar_aba_pagamentos(wb, database):
    """Cria aba de histórico de pagamentos"""
    ws = wb.create_sheet("Pagamentos")
    
    # Headers
    headers = ["ID Pagamento", "ID Empréstimo", "Cliente", "Valor", "Data", "Tipo", "Saldo Anterior"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        EstiloExcel.aplicar_header(cell)
    
    ws.row_dimensions[1].height = 20
    
    # Coletar todos os pagamentos
    todos_pagamentos = []
    for emprestimo in database.emprestimos:
        cliente = database.get_cliente_por_id(emprestimo.cliente_id)
        cliente_nome = cliente.nome if cliente else "N/A"
        
        for pagamento in emprestimo.pagamentos:
            todos_pagamentos.append({
                'id': pagamento['id'],
                'emp_id': emprestimo.id,
                'cliente': cliente_nome,
                'valor': pagamento['valor'],
                'data': pagamento['data'],
                'tipo': pagamento['tipo'],
                'saldo_anterior': pagamento.get('saldo_anterior', 0)
            })
    
    # Ordenar por data (mais recentes primeiro)
    todos_pagamentos.sort(key=lambda x: x['data'], reverse=True)
    
    # Dados
    for row, pgt in enumerate(todos_pagamentos, 2):
        ws.cell(row, 1).value = pgt['id']
        ws.cell(row, 2).value = pgt['emp_id']
        ws.cell(row, 3).value = pgt['cliente']
        ws.cell(row, 4).value = pgt['valor']
        ws.cell(row, 5).value = formatar_data_br(pgt['data'])
        ws.cell(row, 6).value = pgt['tipo']
        ws.cell(row, 7).value = pgt['saldo_anterior']
        
        # Aplicar estilos
        for col in range(1, 4):
            EstiloExcel.aplicar_normal(ws.cell(row, col))
        
        EstiloExcel.aplicar_numerico(ws.cell(row, 4), pgt['valor'])
        EstiloExcel.aplicar_data(ws.cell(row, 5))
        EstiloExcel.aplicar_normal(ws.cell(row, 6))
        EstiloExcel.aplicar_numerico(ws.cell(row, 7), pgt['saldo_anterior'])
    
    # Ajustar colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15


def exportar_apenas_emprestimos(database, caminho_saida=None):
    """Exporta apenas empréstimos com dados completos"""
    if not caminho_saida:
        caminho_saida = f"emprestimos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Empréstimos"
    
    # Headers detalhados
    headers = [
        "ID", "Cliente", "Valor Emprestado", "Taxa Juros (%)", "Prazo (meses)",
        "Valor Total (com juros)", "Juros Totais", "Saldo Devedor", "Total Pago",
        "% Concluído", "Próxima Parcela", "Data Início", "Status", "Observações"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        EstiloExcel.aplicar_header(cell)
    
    ws.row_dimensions[1].height = 25
    
    # Dados
    for row, emprestimo in enumerate(database.emprestimos, 2):
        cliente = database.get_cliente_por_id(emprestimo.cliente_id)
        cliente_nome = cliente.nome if cliente else "N/A"
        
        total_pago = emprestimo.valor_total - emprestimo.saldo_devedor
        percentual_pago = (total_pago / emprestimo.valor_total * 100) if emprestimo.valor_total > 0 else 0
        proxima_parcela = emprestimo.get_proxima_parcela() or "-"
        status = "Quitado" if emprestimo.saldo_devedor <= 0 else "Ativo"
        
        ws.cell(row, 1).value = emprestimo.id
        ws.cell(row, 2).value = cliente_nome
        ws.cell(row, 3).value = emprestimo.valor_emprestado
        ws.cell(row, 4).value = emprestimo.taxa_juros * 100
        ws.cell(row, 5).value = emprestimo.prazo_meses
        ws.cell(row, 6).value = emprestimo.valor_total
        ws.cell(row, 7).value = emprestimo.total_juros
        ws.cell(row, 8).value = emprestimo.saldo_devedor
        ws.cell(row, 9).value = total_pago
        ws.cell(row, 10).value = percentual_pago
        ws.cell(row, 11).value = proxima_parcela
        ws.cell(row, 12).value = formatar_data_br(emprestimo.data_emprestimo)
        ws.cell(row, 13).value = status
        ws.cell(row, 14).value = ""
        
        # Estilos numéricos
        for col in [3, 4, 6, 7, 8, 9]:
            EstiloExcel.aplicar_numerico(ws.cell(row, col), ws.cell(row, col).value)
        
        EstiloExcel.aplicar_numerico(ws.cell(row, 10), percentual_pago)
        ws.cell(row, 10).number_format = '0.00"%"'
        
        EstiloExcel.aplicar_normal(ws.cell(row, 11))
        EstiloExcel.aplicar_data(ws.cell(row, 12))
        EstiloExcel.aplicar_status(ws.cell(row, 13), status)
        EstiloExcel.aplicar_normal(ws.cell(row, 14))
    
    # Ajustar colunas
    larguras = [15, 20, 16, 14, 12, 18, 14, 15, 12, 12, 14, 14, 12, 20]
    for col, largura in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = largura
    
    wb.save(caminho_saida)
    return caminho_saida
